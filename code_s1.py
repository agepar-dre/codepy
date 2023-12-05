import inspect
import re
import traceback
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, NamedStyle
from openpyxl.cell.cell import MergedCell
import funcs_s1 as fxml
import pandas as pd
from datetime import datetime
import numpy as np

def sub_n(dfxc):
    '''Remove espaços desnecessários e substitui '\n' por espaços nas colunas de um DataFrame.'''
    new_columns = []
    for ci in dfxc:
        ci = str(ci)

        if ' ' in ci:
            ci = ci.replace('\n', '')
        else:
            ci = ci.replace('\n', ' ')

        if ci == 'Elegível  QRR':
            ci = 'Elegível QRR'

        new_columns.append(ci)

    return new_columns

def sub_ci(ci):
    '''Remove quebras de linha e reduz múltiplos espaços consecutivos a um único espaço em uma string.'''
    ci = str(ci)

    ci = ci.replace('\n', '')
    
    if '  ' in ci:
        ci = ci.replace('  ', ' ')

    return ci

def letras_para_numeros(texto):
    '''Converte uma representação de coluna de letras para números.'''
    texto = texto.lower()
    valor_total = 0
    
    for letra in texto:
        if 'a' <= letra <= 'z':
            valor_total = valor_total * 26 + (ord(letra) - ord('a') + 1)
    
    return valor_total - 1

def identifica_numero(coluna):
    '''Função que identifica se um valor é um número'''
    for valor in coluna:
        if pd.isna(valor):
            return False
        try:
            float(valor)
            return True
        except:
            return False

def listar_argumentos(funcao):
    '''Definindo uma função para listar os argumentos necessários para chamar uma função'''
    assinatura = inspect.signature(funcao)
    parametros = assinatura.parameters
    argumentos_necessarios = []

    for parametro, info in parametros.items():
        if info.default == inspect.Parameter.empty:
            argumentos_necessarios.append(parametro)

    return argumentos_necessarios

def has_formula(x):
    '''Verifica se uma determinada coluna tem uma fórmula associada em um arquivo de dados.'''
    with open(fr'C:\Users\est.angelo\Documents\codepy10-11\DictDados0.txt', 'r', encoding="UTF-8") as file:
        v = False
        for line in file:
            s = f'Coluna {x} '
            if s in line:
                v = True
        return v

def extract_lines(tdict):
    '''Definindo uma função para extrair linhas de um arquivo e criar um dicionário com os dados'''
    extracting = False

    with open(fr'C:\Users\est.angelo\Documents\codepy10-11\DictDados0.txt', 'r', encoding="UTF-8") as file:
        for line in file:
            if extracting:
                if '------------------------------' in line:
                    break
                l = line.split(': ')
                tdict[l[0]] = sub_ci(l[1])
            elif 'Dicionario de dados' in line:
                extracting = True

with open(fr'C:\Users\est.angelo\Documents\codepy10-11\Relatorio_S1.txt', 'w', encoding="UTF-8") as f:
    f.write('\n')

print_de_debug = ''

array_para_sheet_resumo = [[],[],[],[],[],[]]

# Caminho do arquivo Excel de entrada
input_excel_path = r'C:\Users\est.angelo\Documents\codepy10-11\Simulação_calculos_BRR_v6.xlsx'
xl = pd.ExcelFile(input_excel_path)
nomes_planilhas = xl.sheet_names

d_sheets = {}
    
for i in range(1, len(nomes_planilhas)+1):
    k = f'S{i}'
    n_pls = nomes_planilhas[i-1]
    a_df = pd.read_excel(xl, sheet_name=n_pls)
    out_df = pd.DataFrame(a_df)
    d_sheets[k] = out_df


# Criando um dicionário para armazenar dados extraídos
dict_dados = {}
extract_lines(dict_dados)

a = pd.read_excel(xl, sheet_name=nomes_planilhas[0], header=0)
df = pd.DataFrame(a)

# Encontre as linhas finais que contêm NaN
df = df.dropna(axis=0, how='all')
df = df.dropna(thresh=df.shape[0] - 2)

for colunas_to_float in df.columns:
    try:
        df[colunas_to_float] = df[colunas_to_float].astype(float)
    except:
        df[colunas_to_float] = df[colunas_to_float]

df = df.fillna(0)

df_aux = df.copy()

df.columns = sub_n(df.columns)
df_aux.columns = sub_n(df_aux.columns)

# Criando um dicionário para armazenar informações de funções
dict_func = {}

# Iterando pelos dados extraídos
for i in dict_dados:
    if has_formula(i):
        
        stri = f'fxml.calcular_{i}'
        dfi = listar_argumentos(eval(stri))

        atts = []

        # Criando uma lista de argumentos formatados
        for a in dfi:
            if a.isalpha():
                a = a.lower()
            atts.append(f'df_aux[dict_dados[\'{a}\']]')

        dict_func[i] = atts


workbook = openpyxl.Workbook()
sheet = workbook.active

# Definir o estilo da fonte para a primeira linha (caracteres brancos)
font = Font(color="FFFFFF")  # Branco

# Definir o estilo das bordas para a primeira linha (bordas pretas)
border = Border(left=Side(border_style="thin", color="000000"), 
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"))

# Abrindo um arquivo de saída para escrever resultados
with open(fr'C:\Users\est.angelo\Documents\codepy10-11\Relatorio_S1.txt', 'a', encoding="UTF-8") as f:
    f.write(':\n')
    # Iterando pelos dados extraídos
    for x,y in dict_dados.items():
        col_want = y

        if x in dict_func.items():
            try:
                # Construindo uma string de função e avaliando-a
                attr_str = ', '.join(map(str, dict_func[x]))

                s_sheets_padrao = re.findall('S\d', attr_str)
                if s_sheets_padrao:
                    for ss in s_sheets_padrao:
                        digito = int(ss[1:])
                        attr_str = attr_str.replace(ss, nomes_planilhas[digito-1])
                
                for pl in nomes_planilhas:
                    re_pl = rf'({pl})_(\w+)'
                    # print(re_pl)
                    cor_pl = re.findall(re_pl, attr_str)

                    if cor_pl:
                        for cor in cor_pl:
                            sht, cll = cor
                            var_sub = f"df_aux[dict_dados['{sht}_{cll}']]"
                            sht_df = f"pd.DataFrame(pd.read_excel(xl, sheet_name='{sht}', header=0))"
                            new_var_sub = f'{sht_df}.loc[:, {letras_para_numeros(cll.lower())}]'
                            attr_str = attr_str.replace(var_sub, new_var_sub)

                for i in range(1,9):
                    scol = f'S{i}'

                    if scol in attr_str:
                        scol_true = f"df_aux[dict_dados['{scol}']]"
                        attr_str = attr_str.replace(scol_true,scol)

                func_str = f'np.vectorize(fxml.calcular_{x})({attr_str})'
                    
                print_de_debug = func_str
                
                if col_want in df_aux:
                    df_aux[col_want] = eval(func_str)
                else:
                    # Handle the case where 'Elegível juros' doesn't exist in df_aux
                    print(f"{col_want} not found in df_aux")

            except Exception as exception:
                print(print_de_debug)
                traceback.print_exc()
                f.write(x + ' : ' + 'incompativel com a referencia (exception/error)\n')

                continue

        # Verificando se as colunas são iguais ou têm valores nulos
        if df_aux[col_want].astype(str).equals(df[col_want].astype(str)):
            f.write(x + ' : ' + 'Compativel com a referencia\n')
        elif identifica_numero(df_aux[col_want]) and identifica_numero(df[col_want]):
            if df_aux[col_want].astype(int, errors='ignore').equals(df[col_want].astype(int, errors='ignore')):
                f.write(x + ' : ' + 'Compativel com a referencia\n')
        elif (df[col_want].isnull().sum() + df_aux[col_want].isnull().sum()) == (len(df[col_want]) + len(df_aux[col_want])):
            f.write(x + ' : ' + 'Compativel com a referencia\n')
        else:
            f.write(x + ' : ' + 'incompativel com a referencia\n')

            print('--------------')
            print(df[col_want])
            print(df_aux[col_want])
            print('--------------')

        # Adicionar os dados à primeira coluna não preenchida
        sheet.title = nomes_planilhas[0]
        data = df_aux[col_want].tolist()

        # Adicionar os dados à primeira coluna não preenchida
        column = 1
        while sheet.cell(row=1, column=column).value is not None:
            column += 1

        # Adicionar o cabeçalho da coluna
        sheet.cell(row=1, column=column, value=col_want)

        for i, value in enumerate(data, start=1):
            sheet.cell(row=i+1, column=column, value=value)  # +1 para evitar a sobreposição com o cabeçalho

# Verificar se o estilo já existe
date_style_exists = False
for style_name in workbook.style_names:
    if style_name == 'date_style':
        date_style_exists = True
        break

# Adicionar o estilo se não existir
if not date_style_exists:
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    workbook.add_named_style(date_style)

# Iterar sobre todas as planilhas no workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # Iterar sobre todas as células na planilha
    for row in sheet.iter_rows():
        for cell in row:
            # Verificar se a célula está mesclada
            if isinstance(cell, MergedCell):
                # Para células mescladas, usamos o valor da célula superior esquerda
                cell_value = cell[0][0].value
            else:
                cell_value = cell.value

            # Verificar se o valor da célula é um Timestamp do pandas
            if isinstance(cell_value, pd.Timestamp):
                # Formatar a data para "DD/MM/AAAA"
                formatted_date = cell_value.strftime('%d/%m/%Y')

                # Aplicar o estilo de data à célula
                cell.style = 'date_style'

                # Definir o valor formatado de volta à célula
                cell.value = formatted_date

for cell in sheet[1]:
    cell.fill = PatternFill(start_color="FFD633", end_color="FFD633", fill_type="solid")  # amarelo
    cell.border = border
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

# Definir a cor de preenchimento para todas as células na primeira linha
workbook.save(fr'C:\Users\est.angelo\Documents\codepy10-11\OUTPUT_S1.xlsx')